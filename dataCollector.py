import requests
import urllib.request
from bs4 import BeautifulSoup, SoupStrainer
import re
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet import Worksheet
import time
import random
import os
import urllib3
urllib3.disable_warnings()


def splitJoin(string, seperator, joinWith):
    string = string.split(seperator)
    string = joinWith.join(string)
    return string


def bookTagsUrl(URLList):

    tagLink = []

    for x in URLList:
        tagLink.append("https://www.goodreads.com/book/shelves/" + x[36:])

    return tagLink


def tagExtractor(URL):

    tempTag = {}
    tags = {}
    tagNameList = []
    numberList = []

    r = requests.get(URL)

    soup = BeautifulSoup(r.content, "html.parser")

    data = soup.find_all("div", {"class": "shelfStat"})

    for item in data:

        tagName = item.contents[1].text  # extracts tag name
        tagName = splitJoin(tagName, '-', " ")
        tagName = tagName[1:-1]
        numberOfPeople = item.contents[3].text
        # removes " people" from "28 people" so that only number is left
        number = numberOfPeople[:-8]
        number = splitJoin(number, ',', "")
        number = int(number[1:])
        tempTag.setdefault(tagName, number)
        tags.update(tempTag)

    return tags


def bookInfo(URL):

    r = requests.get(URL)

    soup = BeautifulSoup(r.content, "html.parser")

    descriptionData = soup.find_all("div", {"class": "readable stacked"})
    for item in descriptionData:
        print(item.find_all("span"))


def authorInfo(URL):

    r = requests.get(URL)

    soup = BeautifulSoup(r.content, "html.parser")

    authorData = soup.find_all("h1", {"class": "authorName"})
    for item in authorData:
        author = item.contents[1].text

    return author


def authorBooksURL(URL, booksPerAuthor):
    return str("https://www.goodreads.com/author/list/" + URL[38:] + "?page=1&per_page=" + str(booksPerAuthor) + "&sort=popularity&utf8=✓")


def bookTitleList(URL, booksPerAuthor):

    URL = authorBooksURL(URL, booksPerAuthor)

    titleList = []
    counter = 1

    r = requests.get(URL)
    soup = BeautifulSoup(r.content, "html.parser")

    temp = str(soup.find_all("div", {"class": "mediumTex"}))[
        61:63]  # find the number of books of author
    noOfBooks = int(temp.replace(" ", ""))

    authorBooks = soup.find_all("a", {"class": "bookTitle"})
    for item in authorBooks:
        title = item.contents[1].text
        titleList.append(title)

    return noOfBooks, titleList


def bookURLList(URL, booksPerAuthor):

    URL = authorBooksURL(URL, booksPerAuthor)

    URLList = []
    counter = 1

    r = requests.get(URL)
    soup = BeautifulSoup(r.content, "html.parser")

    for a in soup.find_all('a', href=re.compile('/book/show/')):
        counter += 1
        if counter % 2 == 0:
            link = str("https://www.goodreads.com" + a['href'])
            URLList.append(link)

    return URLList


def checkInternet():
    try:
        response = urllib.request.urlopen('http://www.google.com', timeout=2)
        return True
    except urllib.error.URLError as err:
        return False


def checkInternetInLoop():
    if checkInternet() is True:
        pass
    elif checkInternet() is False:
        print("Internet disconnected")
        time.sleep(5)
        checkInternetInLoop()

http = urllib3.PoolManager()


def getBookSummary(url):
    response = http.request('GET', url)
    soup = BeautifulSoup(response.data, 'html.parser')
    soup = soup.find("div",  {'id': 'descriptionContainer'})
    t = soup.get_text().replace('...more', '').strip()
    if t == '':
        soup = soup.find("span",  {'style': 'display:none'})
        t = soup.get_text()
        if t == '':
            return 'NA'
        else:
            return t
    else:
        return t


def dataCollector(numberOfAuthors=4, booksPerAuthor=3):

    checkInternetInLoop()  # checks internet connection and waits if disconnected

    # data source and criterias
    sourceRead = os.path.join(os.path.dirname(
        __file__), 'data', 'Quotes_goodreads.xlsx')
    sourceWrite = os.path.join(os.path.dirname(
        __file__), 'data', 'Book data.xlsx')
    delayMin = 1  # min delay limit
    delayMax = 2  # max delay limit
    minTag = 500  # minimum number of tags required to save the book

    wbRead = openpyxl.load_workbook(sourceRead)
    wbWrite = openpyxl.load_workbook(sourceWrite)
    # wbWrite= Workbook()
    wsRead = wbRead.get_sheet_by_name("Author data")
    wsWrite = wbWrite.active

    wsWrite['A1'] = 'Author number'
    wsWrite['B1'] = 'Author name'
    wsWrite['C1'] = 'Author URL'
    wsWrite['D1'] = 'Book title'
    wsWrite['E1'] = 'Book gURL'
    wsWrite['F1'] = 'Book tags URL'
    wsWrite['G1'] = 'Book tags'
    wsWrite['H1'] = 'Tag count'
    wsWrite['I1'] = 'Summary'

    print("Scraping has begun")

    row = wsWrite.max_row
    lastAuthor = wsWrite['A' + str(row)].value

    if type(lastAuthor) is str:
        lastAuthor = 0

    for x in range(2 + lastAuthor, 2 + lastAuthor + numberOfAuthors):

        # checks internet connection and waits if disconnected to avoid
        # throwing error
        checkInternetInLoop()

        authorURL = wsRead['C' + str(x)].value
        author = authorInfo(authorURL)
        booksURL = authorBooksURL(authorURL, booksPerAuthor)
        noOfBooks, titleList = bookTitleList(booksURL, booksPerAuthor)
        avaBooksPerAuthor = min(noOfBooks, booksPerAuthor)
        bookLinkList = bookURLList(booksURL, booksPerAuthor)
        tagLinkList = bookTagsUrl(bookLinkList)

        for y in range(0, avaBooksPerAuthor):

            # checks internet connection and waits if disconnected to avoid
            # throwing error
            checkInternetInLoop()

            # delay between scraping of pages to avoid detection
            # time.sleep(random.uniform(delayMin , delayMax)) # adds random

            tags = tagExtractor(tagLinkList[y])
            tagCount = sum(tags.values())

            if tagCount > minTag:
                row = str(row + 1)
                wsWrite['A' + row] = x - 1  # saves author number
                wsWrite['B' + row] = author   # saves author name
                wsWrite['C' + row] = authorURL  # saves author page URL
                wsWrite['D' + row] = titleList[y]  # saves book title
                wsWrite['E' + row] = bookLinkList[y]  # saves book URL
                wsWrite['F' + row] = tagLinkList[y]  # saves tag page URL
                wsWrite['G' + row] = str(tags)  # saves tags
                wsWrite['H' + row] = tagCount  # saves tag count
                # saves summary
                wsWrite['I' + row] = getBookSummary(bookLinkList[y])

                print(x - 1, author, "-",
                      titleList[y], "| Book number:", str(row), " | Tag count:", tagCount)

                row += 1

        wbWrite.save(os.path.join(os.path.dirname(
            __file__), 'data', 'Book data.xlsx'))

        print("Saving data......")

    print("~~~~ Process is complete ~~~~")

dataCollector(300, 10)
